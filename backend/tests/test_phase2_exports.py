import csv
import io
from uuid import uuid4

from openpyxl import load_workbook

from app.api.auth import AuthContext, get_auth
from app.main import app
from app.enums import ResultDecision, ReviewStatus, TaskStatus
from app.models import ScreeningDocumentResult, ScreeningPlan, ScreeningTask, StreamEvent


TASK_TITLE = "任务标题"
RAW_QUERY = "原始问题"
DOCUMENT_TITLE = "文档标题"
AGENT_DECISION = "系统判定"
AGENT_REASON = "系统理由"
MATCHED_CONDITIONS = "命中条件"
REVIEW_DECISION = "复核判定"
REVIEW_NOTE = "复核备注"
REVIEWER_NAME = "复核人"
EVIDENCE_SUMMARY = "证据摘要"


def seeded_export_task(session):
    task = ScreeningTask(
        id=uuid4(),
        owner_id="internal-user",
        title="GPU采购",
        raw_query="哪份合同采购了GPU服务器和存储服务器？",
        status=TaskStatus.completed.value,
        current_stage=TaskStatus.completed.value,
        progress_percent=100,
        metrics={"qmd_result_count": 3},
    )
    session.add(task)
    session.flush()
    session.add(ScreeningPlan(task_id=task.id, plan_json={"conditions": [{"id": "gpu", "description": "采购GPU服务器"}]}))
    session.add(
        ScreeningDocumentResult(
            task_id=task.id,
            document_uri="qmd://contract_docs/equipment-purchase-contract.md",
            document_path="equipment-purchase-contract.md",
            document_title="设备采购合同",
            collection="contract_docs",
            decision=ResultDecision.included.value,
            reason="Agent matched",
            matched_conditions=["gpu"],
            missing_conditions=[],
            evidence=[{"page": 1, "text": "GPU服务器 4台", "source": "qmd", "score": 0.93, "condition_id": "gpu", "artifact_ref": "qmd://contract_docs/equipment-purchase-contract.md"}],
            confidence=0.9,
            review_status=ReviewStatus.reviewed.value,
            review_decision=ResultDecision.included.value,
            review_note="人工确认",
            reviewer_name="张三",
        )
    )
    session.add(StreamEvent(task_id=task.id, sequence=1, event_type="task_created", payload={"task_id": str(task.id)}))
    session.commit()
    return task


def seeded_formula_like_export_task(session):
    task = ScreeningTask(
        id=uuid4(),
        owner_id="internal-user",
        title="=cmd|' /C calc'!A0",
        raw_query="+SUM(1,1)",
        status=TaskStatus.completed.value,
        current_stage=TaskStatus.completed.value,
        progress_percent=100,
        metrics={"qmd_result_count": 1},
    )
    session.add(task)
    session.flush()
    session.add(ScreeningPlan(task_id=task.id, plan_json={"conditions": [{"id": "formula", "description": "formula-like values"}]}))
    session.add(
        ScreeningDocumentResult(
            task_id=task.id,
            document_uri="qmd://contract_docs/formula.md",
            document_path="formula.md",
            document_title="-dangerous title",
            collection="contract_docs",
            decision=ResultDecision.included.value,
            reason="@agent reason",
            matched_conditions=["@condition", "\ttab-condition", "\rcarriage-condition"],
            missing_conditions=[],
            evidence=[{"page": 1, "text": "\n=evidence text", "source": "qmd"}],
            confidence=0.9,
            review_status=ReviewStatus.reviewed.value,
            review_decision=ResultDecision.included.value,
            review_note="\n=unsafe review note",
            reviewer_name="张三",
        )
    )
    session.add(StreamEvent(task_id=task.id, sequence=1, event_type="task_created", payload={"task_id": str(task.id)}))
    session.commit()
    return task


def seeded_empty_failed_export_task(session):
    task = ScreeningTask(
        id=uuid4(),
        owner_id="internal-user",
        title="失败任务",
        raw_query="查询失败原因",
        status=TaskStatus.failed.value,
        current_stage=TaskStatus.failed.value,
        progress_percent=40,
        error_code="qmd_unavailable",
        error_message="QMD service unavailable",
        metrics={"qmd_result_count": 0},
    )
    session.add(task)
    session.add(StreamEvent(task_id=task.id, sequence=1, event_type="task_failed", payload={"task_id": str(task.id)}))
    session.commit()
    return task


def csv_rows(response):
    return list(csv.DictReader(io.StringIO(response.text)))


def xlsx_rows(response):
    workbook = load_workbook(io.BytesIO(response.content), data_only=False)
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    return headers, [{headers[index]: cell for index, cell in enumerate(row)} for row in worksheet.iter_rows(min_row=2)]


def test_export_csv_contains_business_columns(client, db_session):
    session, _ = db_session
    task = seeded_export_task(session)

    response = client.get(f"/api/screening-tasks/{task.id}/export.csv")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    rows = csv_rows(response)
    assert rows[0][TASK_TITLE] == "GPU采购"
    assert rows[0][AGENT_DECISION] == "included"
    assert rows[0][REVIEW_DECISION] == "included"
    assert rows[0][REVIEWER_NAME] == "张三"
    assert "GPU服务器 4台" in rows[0][EVIDENCE_SUMMARY]


def test_export_csv_uses_chinese_headers_and_preserves_empty_task_metadata(client, db_session):
    session, _ = db_session
    task = seeded_empty_failed_export_task(session)

    response = client.get(f"/api/screening-tasks/{task.id}/export.csv")

    assert response.status_code == 200
    rows = csv_rows(response)
    assert TASK_TITLE in rows[0]
    assert rows[0][TASK_TITLE] == "失败任务"
    assert rows[0][RAW_QUERY] == "查询失败原因"
    assert rows[0][DOCUMENT_TITLE] == ""
    assert rows[0][AGENT_DECISION] == ""


def test_export_csv_sanitizes_formula_like_values(client, db_session):
    session, _ = db_session
    task = seeded_formula_like_export_task(session)

    response = client.get(f"/api/screening-tasks/{task.id}/export.csv")

    assert response.status_code == 200
    rows = csv_rows(response)
    row = rows[0]
    assert row[TASK_TITLE] == "'=cmd|' /C calc'!A0"
    assert row[RAW_QUERY] == "'+SUM(1,1)"
    assert row[DOCUMENT_TITLE] == "'-dangerous title"
    assert row[AGENT_REASON] == "'@agent reason"
    assert row[MATCHED_CONDITIONS] == "'@condition, \ttab-condition, \rcarriage-condition"
    assert row[REVIEW_NOTE] == "'\n=unsafe review note"
    assert row[EVIDENCE_SUMMARY] == "'=evidence text"


def test_export_json_contains_task_plan_results_events(client, db_session):
    session, _ = db_session
    task = seeded_export_task(session)

    response = client.get(f"/api/screening-tasks/{task.id}/export.json")

    assert response.status_code == 200
    payload = response.json()
    assert payload["task"]["task_id"] == str(task.id)
    assert payload["task"]["error_code"] is None
    assert payload["task"]["error_message"] is None
    assert payload["plan"]["conditions"][0]["id"] == "gpu"
    assert payload["results"][0]["reviewer_name"] == "张三"
    assert payload["events"][0]["type"] == "task_created"


def test_export_json_keeps_formula_like_values_raw(client, db_session):
    session, _ = db_session
    task = seeded_formula_like_export_task(session)

    response = client.get(f"/api/screening-tasks/{task.id}/export.json")

    assert response.status_code == 200
    payload = response.json()
    assert payload["task"]["title"] == "=cmd|' /C calc'!A0"
    assert payload["task"]["raw_query"] == "+SUM(1,1)"
    assert payload["results"][0]["document_title"] == "-dangerous title"
    assert payload["results"][0]["matched_conditions"] == ["@condition", "\ttab-condition", "\rcarriage-condition"]
    assert payload["results"][0]["review_note"] == "\n=unsafe review note"
    assert payload["results"][0]["evidence"][0]["text"] == "\n=evidence text"


def test_export_json_preserves_failed_task_diagnostics(client, db_session):
    session, _ = db_session
    task = seeded_empty_failed_export_task(session)

    response = client.get(f"/api/screening-tasks/{task.id}/export.json")

    assert response.status_code == 200
    payload = response.json()
    assert payload["task"]["error_code"] == "qmd_unavailable"
    assert payload["task"]["error_message"] == "QMD service unavailable"
    assert payload["results"] == []


def test_export_xlsx_returns_excel_content_type(client, db_session):
    session, _ = db_session
    task = seeded_export_task(session)

    response = client.get(f"/api/screening-tasks/{task.id}/export.xlsx")

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert response.content.startswith(b"PK")


def test_export_xlsx_uses_chinese_headers_and_preserves_empty_task_metadata(client, db_session):
    session, _ = db_session
    task = seeded_empty_failed_export_task(session)

    response = client.get(f"/api/screening-tasks/{task.id}/export.xlsx")

    assert response.status_code == 200
    headers, rows = xlsx_rows(response)
    assert TASK_TITLE in headers
    assert len(rows) == 1
    assert rows[0][TASK_TITLE].value == "失败任务"
    assert rows[0][RAW_QUERY].value == "查询失败原因"
    assert rows[0][DOCUMENT_TITLE].value is None
    assert rows[0][AGENT_DECISION].value is None


def test_export_xlsx_sanitizes_formula_like_values_as_literal_strings(client, db_session):
    session, _ = db_session
    task = seeded_formula_like_export_task(session)

    response = client.get(f"/api/screening-tasks/{task.id}/export.xlsx")

    assert response.status_code == 200
    _, rows = xlsx_rows(response)
    row = rows[0]
    expected_values = {
        TASK_TITLE: "'=cmd|' /C calc'!A0",
        RAW_QUERY: "'+SUM(1,1)",
        DOCUMENT_TITLE: "'-dangerous title",
        AGENT_REASON: "'@agent reason",
        REVIEW_NOTE: "'\n=unsafe review note",
        EVIDENCE_SUMMARY: "'=evidence text",
    }
    for column, expected in expected_values.items():
        assert row[column].value == expected
        assert row[column].data_type == "s"
    assert row[MATCHED_CONDITIONS].value.startswith("'@condition")
    assert row[MATCHED_CONDITIONS].value.replace("\r", "\n") == "'@condition, \ttab-condition, \ncarriage-condition"
    assert row[MATCHED_CONDITIONS].data_type == "s"


def test_export_routes_return_404_for_missing_and_other_owner_tasks(client, db_session):
    session, _ = db_session
    task = seeded_export_task(session)
    missing_id = uuid4()

    missing_response = client.get(f"/api/screening-tasks/{missing_id}/export.json")

    assert missing_response.status_code == 404

    app.dependency_overrides[get_auth] = lambda: AuthContext(owner_id="other-user")
    try:
        other_owner_response = client.get(f"/api/screening-tasks/{task.id}/export.csv")
    finally:
        app.dependency_overrides.pop(get_auth, None)

    assert other_owner_response.status_code == 404
